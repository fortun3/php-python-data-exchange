# -*- coding: utf-8 -*-
"""
Created on Fri Apr  8 11:09:31 2022

"""

import openpyxl

def char(name):
    wb = openpyxl.load_workbook("fruits_Info.xlsx")
    ws = wb['fruit']

    name_list = []
    charname = ws["A"]
    fruit_no = []

    for cell in charname:
        # print(cell.value)
        name_list.append(cell.value)

    # row number = index number + 1
    if name in name_list:
        info_row = name_list.index(name) + 1
        # print(info_row)
        char_info = ws["B"+str(info_row)].value
        # print("")
        # print("Information for " + name + " :")
        # print(char_info)
        if name == "apple":         
            fruit_no.append(1)
            fruit_no.append(2)
            fruit_no.append(4)
        elif name == "banana":         
            fruit_no.append(5)
            fruit_no.append(7)
        elif name == "orange":
            fruit_no.append(3)
            fruit_no.append(6)
            fruit_no.append(8)
        
        # print("")
        # print("displaying all pic of " + name + " : " + str(fruit_no))

        return{
            "description": char_info,
            "pic": fruit_no
        }

    else:
        return{
            "error": "no such fruit"
        }
